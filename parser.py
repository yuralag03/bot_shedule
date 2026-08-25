from openpyxl import load_workbook
import aiosqlite
import re


async def parse_and_save_schedule(file_path: str, group_name: str):
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active

    lessons_to_save = []
    current_day = None
    current_time = None

    # Создаём словарь объединённых ячеек: (row, col) -> верхняя левая ячейка
    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_cells_map[(row, col)] = top_left_cell

    for row_idx in range(5, ws.max_row + 1):
        # Читаем день недели (колонка A)
        day_cell = _get_cell_value(ws, row_idx, 1, merged_cells_map)
        day_val = str(day_cell).strip() if day_cell else ""

        if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
            current_day = day_val

        if not current_day:
            continue

        # Читаем время (колонка B)
        time_cell = _get_cell_value(ws, row_idx, 2, merged_cells_map)
        if time_cell:
            time_str = str(time_cell).strip()
            time_clean = re.sub(r':00', '', time_str).strip()
            current_time = time_clean

        if not current_time:
            continue

        # Читаем предметы для 1 и 2 подгрупп (колонки D и F)
        subj1 = _get_cell_value(ws, row_idx, 4, merged_cells_map)
        subj2 = _get_cell_value(ws, row_idx, 6, merged_cells_map)

        aud1_cell = ws.cell(row=row_idx, column=5)
        aud2_cell = ws.cell(row=row_idx, column=7)
        aud_gen_cell = ws.cell(row=row_idx, column=10)

        # Проверяем, объединены ли ячейки 1 п/г и 2 п/г в один диапазон
        subj1_range = _get_merged_range(ws, row_idx, 4)
        subj2_range = _get_merged_range(ws, row_idx, 6)

        is_merged_across_subgroups = False
        if subj1_range and subj2_range and subj1_range == subj2_range:
            # Ячейки объединены в один диапазон - это общая пара
            is_merged_across_subgroups = True

        # Получаем цвет шрифта из верхней левой ячейки объединённого диапазона
        if is_merged_across_subgroups:
            # Берём цвет из верхней левой ячейки объединённого диапазона
            parity = _get_parity_from_cell(subj1_range.min_row, subj1_range.min_col, ws)
            parity1 = parity
            parity2 = parity
        else:
            # Ячейки не объединены - берём цвет из каждой ячейки отдельно
            parity1 = _get_parity_from_cell(row_idx, 4, ws)
            parity2 = _get_parity_from_cell(row_idx, 6, ws)

        # Обрабатываем 1 подгруппу
        subj1_str = str(subj1).strip() if subj1 else ""
        if subj1_str and subj1_str != "None":
            aud1 = _clean_room(aud1_cell, aud_gen_cell)
            lessons_to_save.append(_parse_lesson(group_name, current_day, current_time, subj1_str, aud1, 1, parity1))

        # Обрабатываем 2 подгруппу
        subj2_str = str(subj2).strip() if subj2 else ""
        if subj2_str and subj2_str != "None":
            aud2 = _clean_room(aud2_cell, aud_gen_cell)
            lessons_to_save.append(_parse_lesson(group_name, current_day, current_time, subj2_str, aud2, 2, parity2))

    wb.close()
    await _save_to_db(group_name, lessons_to_save)
    return len(lessons_to_save)


def _get_cell_value(ws, row, col, merged_cells_map):
    """Получает значение ячейки, учитывая объединённые ячейки"""
    if (row, col) in merged_cells_map:
        return merged_cells_map[(row, col)].value
    return ws.cell(row=row, column=col).value


def _get_merged_range(ws, row, col):
    """Возвращает merged range для ячейки, если она объединена"""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None


def _get_parity_from_cell(row, col, ws):
    """Получает четность недели по цвету шрифта конкретной ячейки"""
    cell = ws.cell(row=row, column=col)
    if not cell.font or not cell.font.color:
        return 0  # Обе недели, если цвет не определен

    color = cell.font.color

    # Проверяем RGB значение
    if color.rgb:
        rgb_str = str(color.rgb).upper()
        # 00333333 - тёмно-серый/чёрный = числитель
        if '00333333' in rgb_str:
            return 1  # Числитель (нечётная неделя)
        # 005871CF - синий = знаменатель
        elif '005871CF' in rgb_str:
            return 2  # Знаменатель (чётная неделя)

    return 0  # Обе недели, если цвет не распознан


def _clean_room(room_cell, gen_room_cell):
    """Очистка аудитории с учетом объединенных ячеек"""
    r1 = str(room_cell.value).strip() if room_cell.value else ""
    r_gen = str(gen_room_cell.value).strip() if gen_room_cell.value else ""

    # Если в ячейке аудитории оказался текст предмета, обнуляем её
    if any(word in r1.lower() for word in ['занятия', 'лекция', 'практические', 'лабораторные', 'не указано',
                                           'экономика', 'философия', 'математика', 'информатика', 'объектно',
                                           'операционные', 'интеграция', 'русский', 'методы']):
        r1 = ""

    # Проверяем, является ли r1 номером аудитории
    if r1 and r1 != "None" and not any(c.isalpha() for c in r1.replace(' ', '').replace('/', '')):
        return r1

    # Иначе берем общую аудиторию
    if r_gen and r_gen != "None" and not any(c.isalpha() for c in r_gen.replace(' ', '').replace('/', '')):
        return r_gen

    return "Не указана"


def _parse_lesson(group_name, day, time, subject_raw, room_raw, subgroup, week_parity):
    parts = [p.strip() for p in str(subject_raw).split('\n') if p.strip()]
    lesson_type = parts[0] if len(parts) > 0 else "Занятие"
    subject_name = parts[1] if len(parts) > 1 else "Не указано"
    teacher = "Не указан"
    if len(parts) > 2 and '(' in parts[2]:
        teacher = parts[2].strip('()')
    room = str(room_raw).strip()

    return {
        'group_name': group_name,
        'day': day,
        'time': time,
        'type': lesson_type,
        'subject': subject_name,
        'teacher': teacher,
        'room': room,
        'subgroup': subgroup,
        'week_parity': week_parity
    }


async def _save_to_db(group_name, lessons):
    async with aiosqlite.connect('bot.db') as db:
        await db.execute("DELETE FROM schedule WHERE group_name = ?", (group_name,))

        await db.executemany('''INSERT INTO schedule 
                                (group_name, day, time, type, subject, teacher, room, subgroup, week_parity) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                             [(l['group_name'], l['day'], l['time'], l['type'], l['subject'],
                               l['teacher'], l['room'], l['subgroup'], l['week_parity']) for l in lessons])
        await db.commit()