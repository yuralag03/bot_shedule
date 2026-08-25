from openpyxl import load_workbook


def debug_colors(file_path):
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active

    print("=== АНАЛИЗ ЦВЕТОВ ШРИФТА ===\n")

    current_day = None

    # Проверяем первые 60 строк (этого достаточно для начала расписания)
    for row_idx in range(1, min(65, ws.max_row + 1)):
        day_cell = ws.cell(row=row_idx, column=1)
        day_val = str(day_cell.value).strip() if day_cell.value else ""

        # Если нашли название дня, запоминаем его
        if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
            current_day = day_val

        # Если день еще не начался или мы за пределами расписания, пропускаем
        if not current_day:
            continue

        # Проверяем колонки с предметами: D (4) для 1 п/г, F (6) для 2 п/г
        for col_idx, col_name in [(4, "1 п/г"), (6, "2 п/г")]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_val = str(cell.value).strip() if cell.value else ""

            # Нас интересуют только ячейки, где реально есть текст предмета
            if cell_val and cell_val != "None":
                print(f"Строка {row_idx}, День: {current_day}, Колонка: {col_name}")
                print(f"  Текст: {repr(cell_val[:60])}")

                font = cell.font
                if font and font.color:
                    print(f"  Цвет шрифта:")
                    print(f"    - type : {font.color.type}")
                    print(f"    - theme: {font.color.theme}")
                    print(f"    - rgb  : {font.color.rgb}")
                    print(f"    - tint : {font.color.tint}")
                else:
                    print("  Цвет шрифта: Не определен (default)")
                print("-" * 50)

    wb.close()
    print("=== КОНЕЦ АНАЛИЗА ===")


if __name__ == "__main__":
    debug_colors('bIPT_252.xlsx')